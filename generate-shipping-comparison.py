import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()

# ── Palette ──────────────────────────────────────────────
GOLD     = "C9A870"
DARK     = "1A1A1A"
LIGHT_BG = "FFF8EE"
SUBTOT   = "F5F2EC"
GREEN_BG = "E8F7EB"
BLUE_BG  = "EBF3FF"
YELLOW   = "FFFDE7"
WHITE    = "FFFFFF"
HDR_BG   = "2C2C2C"

ILS  = '₪#,##0'
PCT  = '0%'
NUM0 = '#,##0'

def fnt(bold=False, size=11, color="1A1A1A", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def fill(c):
    return PatternFill("solid", fgColor=c)

def al(wrap=False, h="right"):
    return Alignment(horizontal=h, vertical="center", readingOrder=2, wrap_text=wrap)

def set_rtl(ws):
    ws.sheet_view.rightToLeft = True

def title_row(ws, row, text, cols):
    c = ws.cell(row=row, column=1, value=text)
    c.font = fnt(bold=True, size=14, color=WHITE)
    c.fill = fill(DARK)
    c.alignment = al()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 28

def sec(ws, row, text, cols):
    c = ws.cell(row=row, column=1, value=text)
    c.font = fnt(bold=True, size=10, color=GOLD)
    c.fill = fill(DARK)
    c.alignment = al()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 20

def hdr(ws, row, col, text, wrap=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = fnt(bold=True, color=WHITE, size=10)
    c.fill = fill(HDR_BG)
    c.alignment = al(wrap=wrap)
    ws.row_dimensions[row].height = 36

def inp(ws, row, col, bg=YELLOW):
    c = ws.cell(row=row, column=col, value="")
    c.fill = fill(bg)
    c.alignment = al()
    return c

def lbl(ws, row, col, text, bold=False, color="1A1A1A", bg=None, italic=False, size=11):
    c = ws.cell(row=row, column=col, value=text)
    c.font = fnt(bold=bold, color=color, italic=italic, size=size)
    c.alignment = al()
    if bg:
        c.fill = fill(bg)
    return c

# ══════════════════════════════════════════════════════════
# SHEET 1 — השוואת חברות
# ══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "השוואת חברות"
set_rtl(ws1)

TOTAL_COLS = 24

# Column widths
ws1.column_dimensions["A"].width = 30   # שם חברה
ws1.column_dimensions["B"].width = 13   # מחיר בסיס
ws1.column_dimensions["C"].width = 13   # חריג אורך
ws1.column_dimensions["D"].width = 13   # פריפריה
ws1.column_dimensions["E"].width = 11   # היטל דלק
ws1.column_dimensions["F"].width = 13   # ניסיון מסירה
ws1.column_dimensions["G"].width = 15   # החזרה
ws1.column_dimensions["H"].width = 13   # זמן מרכז
ws1.column_dimensions["I"].width = 13   # זמן פריפריה
ws1.column_dimensions["J"].width = 14   # תדירות איסוף
ws1.column_dimensions["K"].width = 12   # cut-off
ws1.column_dimensions["L"].width = 13   # מינימום
ws1.column_dimensions["M"].width = 14   # ביטוח
ws1.column_dimensions["N"].width = 13   # השתתפות עצמית
ws1.column_dimensions["O"].width = 16   # דרישות אריזה
ws1.column_dimensions["P"].width = 16   # SLA תביעות
ws1.column_dimensions["Q"].width = 11   # API
ws1.column_dimensions["R"].width = 15   # מדפסת
ws1.column_dimensions["S"].width = 18   # איש קשר
ws1.column_dimensions["T"].width = 16   # תקופת חסד [NEW]
ws1.column_dimensions["U"].width = 18   # חבילה בודדת [NEW]
ws1.column_dimensions["V"].width = 18   # אופן קריאת איסוף [NEW]
ws1.column_dimensions["W"].width = 14   # טלפון
ws1.column_dimensions["X"].width = 16   # תנאי תשלום

# Row 1: Title
title_row(ws1, 1, "SPINZ — השוואת הצעות מחיר שליחויות", TOTAL_COLS)

# Row 2: subtitle
ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
c = ws1.cell(row=2, column=1, value="חבילה: 130×18×74 ס\"מ · 12 ק\"ג · מהמחסן עד הלקוח")
c.font = fnt(italic=True, color="888888", size=10)
c.alignment = al()
ws1.row_dimensions[2].height = 16

ws1.row_dimensions[3].height = 6

# Row 4: Section group labels
for col in range(1, TOTAL_COLS + 1):
    ws1.cell(row=4, column=col).fill = fill(DARK)

for start, end, label in [
    (1,  7,  "── עלויות ──"),
    (8,  12, "── שירות ──"),
    (13, 16, "── ביטוח ──"),
    (17, 22, "── תפעול & קשר ──"),
    (23, 24, "── פיננסי ──"),
]:
    ws1.merge_cells(start_row=4, start_column=start, end_row=4, end_column=end)
    c = ws1.cell(row=4, column=start, value=label)
    c.font = fnt(bold=True, color=GOLD, size=10)
    c.fill = fill(DARK)
    c.alignment = al(h="center")
ws1.row_dimensions[4].height = 20

# Row 5: Column headers
headers = [
    "שם חברה",
    "מחיר בסיס (₪)",
    "תוספת חריג אורך (₪)",
    "תוספת פריפריה (₪)",
    "היטל דלק (%)",
    "ניסיון מסירה נוסף (₪)",
    "החזרה / לוגיסטיקה הפוכה (₪)",
    "זמן אספקה מרכז (ימים)",
    "זמן אספקה פריפריה (ימים)",
    "תדירות איסוף",
    "Cut-off time",
    "מינימום חבילות לאיסוף",
    "ביטוח סטנדרטי - מכסה עד (₪)",
    "השתתפות עצמית (₪)",
    "דרישות אריזה למימוש ביטוח",
    "SLA תביעות נזק - כמה זמן לזיכוי?",
    "API / אינטגרציה",
    "מדפסת טרמית - מי מספק?",
    "איש קשר אישי",
    "תקופת חסד למינימום (חודשים)",
    "מדיניות חבילה בודדת / איסוף ריק",
    "אופן פתיחת קריאת איסוף",
    "טלפון ישיר",
    "תנאי תשלום (שוטף +?)",
]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=5, column=col, value=h)
    c.font = fnt(bold=True, color=WHITE, size=10)
    c.fill = fill(HDR_BG)
    c.alignment = al(wrap=True)
ws1.row_dimensions[5].height = 40

# Companies list (no duplicates, all 14)
companies = [
    "כץ משלוחים (Katz)",
    "HFD שליחויות ולוגיסטיקה",
    "קארגו (Kargo)",
    "צ'יטה שליחויות (Cheetah)",
    "תפוז שליחויות (Tapuz)",
    "YDM",
    "דואר ישראל (דואר שליחים)",
    "הום דילברי (Home Delivery)",
    "יונט קארגו",
    "נאוי לוגיסטיקה",
    "ATA לוגיסטיקה",
    "Box.it",
    "Speedex",
    "TNT / FedEx",
]

PURPLE = "EDE7F6"

# Section bg colors — two variants per column (even/odd rows)
col_bg = {
    1:  ("FFFFFF", "F0EDE8"),
    2:  (LIGHT_BG, "FFE8C0"), 3:  (LIGHT_BG, "FFE8C0"), 4:  (LIGHT_BG, "FFE8C0"),
    5:  (LIGHT_BG, "FFE8C0"), 6:  (LIGHT_BG, "FFE8C0"), 7:  (LIGHT_BG, "FFE8C0"),
    8:  (BLUE_BG, "D5E5F5"), 9:  (BLUE_BG, "D5E5F5"), 10: (BLUE_BG, "D5E5F5"),
    11: (BLUE_BG, "D5E5F5"), 12: (BLUE_BG, "D5E5F5"),
    13: (GREEN_BG, "C8E0CC"), 14: (GREEN_BG, "C8E0CC"), 15: (GREEN_BG, "C8E0CC"),
    16: (GREEN_BG, "C8E0CC"),
    17: (SUBTOT, "E4E0D8"), 18: (SUBTOT, "E4E0D8"), 19: (SUBTOT, "E4E0D8"),
    20: (SUBTOT, "E4E0D8"), 21: (SUBTOT, "E4E0D8"), 22: (SUBTOT, "E4E0D8"),
    23: (PURPLE, "D9CFF0"), 24: (PURPLE, "D9CFF0"),
}

for i, company in enumerate(companies):
    row = 6 + i
    even = (i % 2 == 0)

    # Company name
    c = ws1.cell(row=row, column=1, value=company)
    c.font = fnt(bold=True, size=11)
    c.alignment = al()
    c.fill = fill(col_bg[1][0] if even else col_bg[1][1])

    # Input cells
    for col in range(2, TOTAL_COLS + 1):
        bg = col_bg.get(col, (YELLOW, YELLOW))
        cell = ws1.cell(row=row, column=col, value="")
        cell.fill = fill(bg[0] if even else bg[1])
        cell.alignment = al()

    ws1.row_dimensions[row].height = 20

# Freeze panes: freeze row 5 headers + column A
ws1.freeze_panes = "B6"


# ══════════════════════════════════════════════════════════
# SHEET 2 — סימולטור עלות חודשית
# ══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("סימולטור עלות חודשית")
set_rtl(ws2)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 22
ws2.column_dimensions["E"].width = 18

title_row(ws2, 1, "SPINZ — סימולטור עלות שליחויות חודשית", 5)

ws2.merge_cells("A2:E2")
c = ws2.cell(row=2, column=1, value="שנה את כמות החבילות — העלות החודשית מחושבת אוטומטית")
c.font = fnt(italic=True, color="888888", size=10)
c.alignment = al()
ws2.row_dimensions[2].height = 16
ws2.row_dimensions[3].height = 6

# Input: כמות חבילות
sec(ws2, 4, "קלט", 5)
lbl(ws2, 5, 1, "כמות חבילות בחודש:", bold=True)
qty = ws2.cell(row=5, column=2, value=50)
qty.font = fnt(bold=True, size=14, color=GOLD)
qty.fill = fill(YELLOW)
qty.alignment = al()
qty.number_format = NUM0
lbl(ws2, 5, 3, "← שנה כאן", italic=True, color="888888")
ws2.row_dimensions[5].height = 26
ws2.row_dimensions[6].height = 6

# Headers
sec(ws2, 7, "תוצאות", 5)
for col, h in enumerate(["שם חברה", "מחיר בסיס (₪)", "תוספות ממוצע (₪)", "עלות לחבילה (₪)", "עלות חודשית (₪)"], 1):
    hdr(ws2, 8, col, h, wrap=False)

# Company rows
for i, company in enumerate(companies):
    row = 9 + i
    row_bg = SUBTOT if i % 2 == 0 else None

    c = ws2.cell(row=row, column=1, value=company)
    c.font = fnt(bold=True)
    c.alignment = al()
    if row_bg:
        c.fill = fill(row_bg)

    for col in [2, 3]:
        cell = ws2.cell(row=row, column=col, value="")
        cell.fill = fill(YELLOW if not row_bg else "FFF0D4")
        cell.alignment = al()
        cell.number_format = ILS

    # עלות לחבילה = B+C
    f1 = ws2.cell(row=row, column=4,
                  value=f"=IF(AND(B{row}<>\"\",C{row}<>\"\"),B{row}+C{row},\"\")")
    f1.font = fnt(bold=True)
    f1.alignment = al()
    f1.number_format = ILS
    if row_bg:
        f1.fill = fill(row_bg)

    # עלות חודשית = D × כמות
    f2 = ws2.cell(row=row, column=5,
                  value=f"=IF(D{row}<>\"\",D{row}*$B$5,\"\")")
    f2.font = fnt(bold=True, color=GOLD)
    f2.alignment = al()
    f2.number_format = ILS
    if row_bg:
        f2.fill = fill(row_bg)
    else:
        f2.fill = fill(LIGHT_BG)

    ws2.row_dimensions[row].height = 20

# Totals
total_row = 9 + len(companies)
ws2.row_dimensions[total_row].height = 6
summary_row = total_row + 1

for col in range(1, 6):
    ws2.cell(row=summary_row, column=col).fill = fill(DARK)

ws2.merge_cells(f"A{summary_row}:D{summary_row}")
c = ws2.cell(row=summary_row, column=1, value="הזולה ביותר (חודשי)")
c.font = fnt(bold=True, color=WHITE)
c.fill = fill(DARK)
c.alignment = al()

f_min = ws2.cell(row=summary_row, column=5,
                 value=f"=IFERROR(MIN(E9:E{total_row-1}),\"\")")
f_min.font = fnt(bold=True, color=GOLD)
f_min.fill = fill(DARK)
f_min.alignment = al()
f_min.number_format = ILS
ws2.row_dimensions[summary_row].height = 22

ws2.freeze_panes = "A9"


# ══════════════════════════════════════════════════════════
# SHEET 3 — שאלות לפנייה לחברות
# ══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("שאלות לחברות")
set_rtl(ws3)

ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 55
ws3.column_dimensions["C"].width = 36

title_row(ws3, 1, "SPINZ — שאלות לשאול כל חברת שליחויות", 3)

ws3.merge_cells("A2:C2")
c = ws3.cell(row=2, column=1, value="העתק ושלח בפנייה לכל חברה — בקש מענה בכתב")
c.font = fnt(italic=True, color="888888", size=10)
c.alignment = al()
ws3.row_dimensions[2].height = 16
ws3.row_dimensions[3].height = 6

for col, h in enumerate(["נושא", "שאלה", "תשובה"], 1):
    hdr(ws3, 4, col, h, wrap=False)

questions = [
    ("עלויות",    "מה המחיר לחבילה בגודל 130×18×74 ס\"מ / 12 ק\"ג?"),
    ("עלויות",    "האם יש תוספת חריג על אורך מעל מטר? כמה?"),
    ("עלויות",    "האם יש תוספת לפריפריה / יישובי קצה?"),
    ("עלויות",    "האם המחיר כפוף להיטל דלק? מה האחוז הנוכחי?"),
    ("עלויות",    "מה עלות ניסיון מסירה שני/שלישי כשהלקוח לא זמין?"),
    ("עלויות",    "מה עלות לוגיסטיקה הפוכה (החזרה מלקוח)?"),
    ("שירות",     "מה זמן האספקה ללקוח — מרכז הארץ?"),
    ("שירות",     "מה זמן האספקה — פריפריה / ערים רחוקות?"),
    ("שירות",     "מה תדירות האיסוף מהמחסן שלנו? (יומי / יומיים / שבועי)"),
    ("שירות",     "מה שעת ה-cut-off היומית לאיסוף?"),
    ("שירות",     "האם יש מינימום כמות חבילות לאיסוף?"),
    ("ביטוח",     "מה גובה הביטוח הסטנדרטי? מה מכוסה?"),
    ("ביטוח",     "מה גובה ההשתתפות העצמית במקרה נזק / אובדן?"),
    ("ביטוח",     "האם יש דרישות אריזה ספציפיות למימוש הביטוח?"),
    ("תפעול",     "האם יש API או פלאגין לחיבור לאתר?"),
    ("תפעול",     "האם אתם מספקים מדפסת טרמית (Zebra וכד') וגלילים?"),
    ("תפעול",     "כעסק חדש בהשקה — האם נוכל לקבל תקופת חסד של 6 חודשים ללא התניית מינימום איסוף יומי?"),
    ("תפעול",     "במידה וביום מסוים יש רק חבילה אחת (או אפס) — האם המשאית מגיעה כרגיל, או שההזמנה נדחית ליום הבא? האם יש חיוב על הגעת סרק?"),
    ("תפעול",     "איך מתבצע תיאום האיסוף היומי בפועל? (API אוטומטי / פורטל החברה / טלפון/וואטסאפ לסדרן)"),
    ("קשר",       "האם יש מנהל תיק אישי עם טלפון ישיר (לא מוקד)?"),
    ("פיננסי",    "מהם תנאי התשלום? (שוטף+30 / שוטף+60 / כרטיס אשראי מתגלגל)"),
    ("פיננסי",    "במידה ומופעל ביטוח על נזק — תוך כמה זמן מבוצע הזיכוי בפועל?"),
]

cat_bg = {
    "עלויות":  LIGHT_BG,
    "שירות":   BLUE_BG,
    "ביטוח":   GREEN_BG,
    "תפעול":   YELLOW,
    "קשר":     SUBTOT,
    "פיננסי":  "EDE7F6",
}

for i, (cat, q) in enumerate(questions):
    row = 5 + i
    bg = cat_bg.get(cat, SUBTOT)

    c1 = ws3.cell(row=row, column=1, value=cat)
    c1.font = fnt(bold=True, size=10)
    c1.fill = fill(bg)
    c1.alignment = al()

    c2 = ws3.cell(row=row, column=2, value=q)
    c2.font = fnt(size=11)
    c2.alignment = al(wrap=True)

    c3 = ws3.cell(row=row, column=3, value="")
    c3.fill = fill(YELLOW)
    c3.alignment = al(wrap=True)

    ws3.row_dimensions[row].height = 22

# ── Save ──────────────────────────────────────────────────
out = "spinz-shipping-comparison.xlsx"
wb.save(out)
sys.stdout.write("Done: " + out + "\n")
