import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Palette ──────────────────────────────────────────────
GOLD      = "C9A870"
DARK      = "1A1A1A"
LIGHT_BG  = "FFF8EE"
SUBTOT_BG = "F5F2EC"
GREEN_BG  = "E8F7EB"
WARN_BG   = "FFF3CD"
BLUE_BG   = "EBF3FF"
WHITE     = "FFFFFF"
HDR_BG    = "2C2C2C"
INPUT_BG  = "FFFDE7"   # yellow = editable

def fnt(bold=False, size=11, color="1A1A1A", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def fill(c):
    return PatternFill("solid", fgColor=c)

def al(wrap=False):
    return Alignment(horizontal="right", vertical="center",
                     readingOrder=2, wrap_text=wrap)

def set_rtl(ws):
    ws.sheet_view.rightToLeft = True

def title_row(ws, row, text, cols=5):
    c = ws.cell(row=row, column=1, value=text)
    c.font = fnt(bold=True, size=14, color=WHITE)
    c.fill = fill(DARK)
    c.alignment = al()
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 28

def sec(ws, row, text, cols=5):
    c = ws.cell(row=row, column=1, value=text)
    c.font = fnt(bold=True, size=10, color=GOLD)
    c.fill = fill(DARK)
    c.alignment = al()
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 20

def hdr(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = fnt(bold=True, color=WHITE)
    c.fill = fill(HDR_BG)
    c.alignment = al()

def inp(ws, row, col, value, fmt=None, note_col=None, note=None):
    """Yellow input cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = fnt(bold=True, color=GOLD)
    c.fill = fill(INPUT_BG)
    c.alignment = al()
    if fmt: c.number_format = fmt
    if note_col and note:
        n = ws.cell(row=row, column=note_col, value=note)
        n.font = fnt(italic=True, color="888888", size=9)
        n.alignment = al()

def lbl(ws, row, col, text, bold=False, color="1A1A1A", bg=None, italic=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = fnt(bold=bold, color=color, italic=italic)
    c.alignment = al()
    if bg: c.fill = fill(bg)

def val(ws, row, col, formula, fmt=None, bold=False, bg=None, color="1A1A1A"):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = fnt(bold=bold, color=color)
    c.alignment = al()
    if fmt: c.number_format = fmt
    if bg: c.fill = fill(bg)

ILS  = '₪#,##0'
USD  = '"$"#,##0.00'
PCT  = '0.0%'
NUM  = '#,##0.0'
NUM0 = '#,##0'

S1 = "'מודל תמחור (20 מול 40 פיט)'"

# ══════════════════════════════════════════════════════════
# SHEET 1 — מודל תמחור
# ══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "מודל תמחור (20 מול 40 פיט)"
set_rtl(ws1)
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 32
ws1.row_dimensions[1].height = 28

title_row(ws1, 1, "SPINZ — מודל תמחור ייבוא", cols=4)
ws1.row_dimensions[2].height = 6
sec(ws1, 3, "פרמטרים", cols=4)

# Params — B4:B6 are the anchor cells
params = [
    (4, "שער הגנה (Buffer Rate) ₪/$", 3.15, '₪0.00',  "שמרני — לתכנון קדימה בלבד"),
    (5, 'מע"מ',                        0.17, PCT,       "17% — מוחזר ~60 יום"),
    (6, 'בצ"מ (בלתי צפוי)',             0.05, PCT,       "כרית ביטחון 5%"),
]
for row, label, value, fmt, note in params:
    lbl(ws1, row, 1, label, bold=True)
    inp(ws1, row, 2, value, fmt=fmt, note_col=4, note=note)
    ws1.row_dimensions[row].height = 18

ws1.row_dimensions[7].height = 6
sec(ws1, 8, "פירוט עלות ליחידה", cols=4)

for col, txt in [(1,"סעיף"),(2,"מכולה 20 פיט (170 יח')"),(3,"מכולה 40 פיט (330 יח')"),(4,"הערות")]:
    hdr(ws1, 9, col, txt)
ws1.row_dimensions[9].height = 20

# Raw inputs (USD)
inputs_data = [
    (10, "מחיר מוצר (FOB) $",             93.80, 93.80, "עד נמל סין"),
    (11, "שילוח ימי $",                    16.06, 11.50, "חלק יחסי מעלות המכולה"),
    (12, "לוגיסטיקה יבשתית $ (worst case)",12.00, 12.00, "עמיל מכס + דמי נמל + ביטוח + שילוח פנים"),
]
for row, label, b, c, note in inputs_data:
    lbl(ws1, row, 1, label)
    ws1.cell(row=row, column=2, value=b).number_format = USD
    ws1.cell(row=row, column=2).alignment = al()
    ws1.cell(row=row, column=3, value=c).number_format = USD
    ws1.cell(row=row, column=3).alignment = al()
    n = ws1.cell(row=row, column=4, value=note)
    n.font = fnt(italic=True, color="888888", size=9)
    n.alignment = al(wrap=True)
    ws1.row_dimensions[row].height = 18

# Formula rows
formula_rows = [
    (13, "סך עלות בדולרים $",
         "=B10+B11+B12", "=C10+C11+C12", USD, True, SUBTOT_BG, ""),
    (14, "סך עלות בשקלים (× שער הגנה)",
         "=B13*$B$4", "=C13*$B$4", ILS, False, None, ""),
    (15, 'תוספת בצ"מ 5%',
         "=B14*$B$6", "=C14*$B$6", ILS, False, None, ""),
    (16, "עלות נחיתה לפני מע\"מ ₪",
         "=B14+B15", "=C14+C15", ILS, True, SUBTOT_BG, "הבסיס לחישוב רווחיות"),
    (17, 'מע"מ 17% ₪  — cash flow בלבד',
         "=B16*$B$5", "=C16*$B$5", ILS, False, LIGHT_BG, "מוחזר ~60 יום — אינו עלות"),
    (18, "עלות כוללת מהכיס ₪ (כולל מע\"מ)",
         "=B16+B17", "=C16+C17", ILS, True, DARK, ""),
]
for row, label, fb, fc, fmt, bold, bg, note in formula_rows:
    color = WHITE if bg == DARK else "1A1A1A"
    lbl(ws1, row, 1, label, bold=bold, color=color, bg=bg,
        italic=(bg==LIGHT_BG))
    val(ws1, row, 2, fb, fmt=fmt, bold=bold, bg=bg, color=color)
    val(ws1, row, 3, fc, fmt=fmt, bold=bold, bg=bg, color=color)
    if note:
        n = ws1.cell(row=row, column=4, value=note)
        n.font = fnt(italic=True, color="888888", size=9)
        n.alignment = al()
    ws1.row_dimensions[row].height = 20 if bold else 18

# ══════════════════════════════════════════════════════════
# SHEET 2 — סימולטור
# ══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("סימולטור משחק (דמו)")
set_rtl(ws2)
ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 20
ws2.column_dimensions["D"].width = 20
ws2.column_dimensions["E"].width = 18

title_row(ws2, 1, "SPINZ — סימולטור כדאיות & Break-even", cols=5)
ws2.row_dimensions[2].height = 6

# ── Section A: Inputs ──────────────────────────────────
sec(ws2, 3, "קלטים", cols=5)

for col, txt in [(1,"פרמטר"),(2,"ערך"),(3,"הערה")]:
    hdr(ws2, 4, col, txt)
ws2.row_dimensions[4].height = 18

lbl(ws2, 5, 1, "בחירת מכולה (הקלד 20 או 40)", bold=True)
inp(ws2, 5, 2, 20, fmt=NUM0, note_col=3, note="20 = מכולת 20 פיט (170 יח') | 40 = מכולת 40 פיט (330 יח')")

# Container validation
dv = DataValidation(type="whole", operator="between", formula1="20", formula2="40")
dv.sqref = "B5"
ws2.add_data_validation(dv)

lbl(ws2, 6, 1, "שער חליפין (חי / לבדיקה) ₪/$", bold=True)
inp(ws2, 6, 2, 3.65, fmt='₪0.00', note_col=3, note="שחק עם השער — לא משפיע על גיליון 1")

lbl(ws2, 7, 1, "מחיר יחידה לצרכן (כולל מע\"מ) ₪", bold=True)
inp(ws2, 7, 2, 1499, fmt=ILS, note_col=3, note="המחיר שתציג ללקוח")

lbl(ws2, 8, 1, "עלות שיווק ליחידה (CPA) ₪", bold=True)
inp(ws2, 8, 2, 80, fmt=ILS, note_col=3, note="פרסום + עמלות + קידום")

ws2.row_dimensions[9].height = 6

# ── Section B: Results ────────────────────────────────
sec(ws2, 10, "תוצאות", cols=5)

# Landing cost formula: picks 20ft or 40ft from sheet1, uses simulator rate
# pre-VAT (B16 or C16 from sheet1 based on container), converted back to USD then * sim rate
# Actually: raw USD cost * sim rate * (1 + contingency)
land_usd_20 = f"({S1}!B10+{S1}!B11+{S1}!B12)"
land_usd_40 = f"({S1}!C10+{S1}!C11+{S1}!C12)"
landing_formula = f"=IF(B5=20,{land_usd_20},{land_usd_40})*B6*(1+{S1}!B6)"

results = [
    (11, "עלות נחיתה לפני מע\"מ ₪ (לפי שער הסימולטור)",
         landing_formula, ILS, True, SUBTOT_BG,
         "בסיס לחישוב רווח — מע\"מ מוחזר ואינו עלות אמיתית"),
    (12, "מחיר לצרכן ללא מע\"מ ₪",
         "=B7/1.17", ILS, False, None, ""),
    (13, "רווח גולמי ₪",
         "=B12-B11-B8", ILS, True, GREEN_BG, ""),
    (14, "אחוז רווחיות (Margin)",
         "=IF(B12>0,B13/B12,0)", PCT, True, GREEN_BG,
         "מעל 30% טוב | מעל 40% מצוין"),
]
for row, label, formula, fmt, bold, bg, note in results:
    color = "1A1A1A"
    lbl(ws2, row, 1, label, bold=bold, bg=bg)
    val(ws2, row, 2, formula, fmt=fmt, bold=bold, bg=bg)
    if note:
        n = ws2.cell(row=row, column=3, value=note)
        n.font = fnt(italic=True, color="888888", size=9)
        n.alignment = al()
    ws2.row_dimensions[row].height = 20 if bold else 18

ws2.row_dimensions[15].height = 6

# ── Section C: Break-even ────────────────────────────
sec(ws2, 16, "Break-even — כמה זמן לאחזיר כל הוצאה?", cols=5)

note_be = ("הרווח ליחידה נמשך אוטומטית מהתוצאות למעלה. "
           "הזן סכום ₪ בעמודה B — שאר העמודות מחושבות.")
ws2.merge_cells("A17:E17")
c = ws2.cell(row=17, column=1, value=note_be)
c.font = fnt(italic=True, color="888888", size=9)
c.alignment = al(wrap=True)
ws2.row_dimensions[17].height = 28

for col, txt in [(1,"הוצאה"),(2,"סכום ₪"),(3,"יחידות לכיסוי"),(4,"מנות 20 פיט"),(5,"מנות 40 פיט")]:
    hdr(ws2, 18, col, txt)
ws2.row_dimensions[18].height = 20

be_items = [
    "טיסות לסין",
    "בדיקות תקן / אישורים",
    "צילומים ותוכן",
    "בניית עמוד / אתר",
    "משפיענים (launch)",
    "קמפיין אינסטגרם",
    "הוצאה נוספת",
    "הוצאה נוספת",
]

for i, item in enumerate(be_items):
    row = 19 + i
    bg = SUBTOT_BG if i % 2 == 0 else None

    # Label
    lbl(ws2, row, 1, item, bg=bg)

    # Amount — yellow input
    c = ws2.cell(row=row, column=2, value=0)
    c.font = fnt(bold=True, color=GOLD)
    c.fill = fill(INPUT_BG)
    c.alignment = al()
    c.number_format = ILS

    # Units to break-even = amount / gross profit per unit
    # Guard against divide-by-zero
    f_units = f"=IF(B13>0,CEILING(B{row}/B13,1),\"—\")"
    c3 = ws2.cell(row=row, column=3, value=f_units)
    c3.font = fnt(bold=True)
    c3.alignment = al()
    c3.number_format = NUM0
    if bg: c3.fill = fill(bg)

    # Batches of 20ft
    f_b20 = f"=IF(ISNUMBER(C{row}),CEILING(C{row}/170,0.1),\"—\")"
    c4 = ws2.cell(row=row, column=4, value=f_b20)
    c4.font = fnt()
    c4.alignment = al()
    c4.number_format = NUM
    if bg: c4.fill = fill(bg)

    # Batches of 40ft
    f_b40 = f"=IF(ISNUMBER(C{row}),CEILING(C{row}/330,0.1),\"—\")"
    c5 = ws2.cell(row=row, column=5, value=f_b40)
    c5.font = fnt()
    c5.alignment = al()
    c5.number_format = NUM
    if bg: c5.fill = fill(bg)

    ws2.row_dimensions[row].height = 18

# Totals row
total_row = 19 + len(be_items)
lbl(ws2, total_row, 1, "סה\"כ השקעה ראשונית", bold=True, bg=DARK, color=WHITE)
val(ws2, total_row, 2, f"=SUM(B19:B{total_row-1})", fmt=ILS, bold=True, bg=DARK, color=WHITE)
val(ws2, total_row, 3, f"=IF(B13>0,CEILING(B{total_row}/B13,1),\"—\")", fmt=NUM0, bold=True, bg=DARK, color=WHITE)
val(ws2, total_row, 4, f"=IF(ISNUMBER(C{total_row}),CEILING(C{total_row}/170,0.1),\"—\")", fmt=NUM, bold=True, bg=DARK, color=WHITE)
val(ws2, total_row, 5, f"=IF(ISNUMBER(C{total_row}),CEILING(C{total_row}/330,0.1),\"—\")", fmt=NUM, bold=True, bg=DARK, color=WHITE)
ws2.row_dimensions[total_row].height = 22

# ══════════════════════════════════════════════════════════
# SHEET 3 — תזרים בפועל
# ══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("תזרים והוצאות בפועל")
set_rtl(ws3)
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 36
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 20
ws3.column_dimensions["E"].width = 20

title_row(ws3, 1, "SPINZ — תזרים תשלומים בפועל (היסטורי)", cols=5)

ws3.merge_cells("A2:E2")
c = ws3.cell(row=2, column=1,
    value="גיליון היסטורי — שער ההמרה מוקלד ידנית. אין עדכון אוטומטי.")
c.font = fnt(italic=True, color="C0392B", size=9)
c.alignment = al()
c.fill = fill(WARN_BG)
ws3.row_dimensions[2].height = 16

for col, txt in enumerate(
    ["תאריך", "תיאור העסקה", 'סכום מט"ח (USD)', "שער המרה בפועל", "סכום בשקלים ₪"], 1):
    hdr(ws3, 3, col, txt)
ws3.row_dimensions[3].height = 20

samples = [
    ("2025-04-10", "מקדמה 30% לספק בסין",           4793.4),
    ("2025-05-01", "תשלום יתרה 70% לספק",            11185.1),
    ("2025-05-20", "תשלום שילוח ימי — מכולה 20 פיט",  2730.2),
    ("2025-06-05", "תשלום לוגיסטיקה (עמיל + נמל)",    2040.0),
]
for i, (date, desc, usd) in enumerate(samples, 4):
    ws3.cell(row=i, column=1, value=date).alignment = al()
    ws3.cell(row=i, column=1).font = fnt(color="555555")
    ws3.cell(row=i, column=2, value=desc).alignment = al()
    c = ws3.cell(row=i, column=3, value=usd)
    c.number_format = USD; c.alignment = al()
    d = ws3.cell(row=i, column=4)
    d.fill = fill(INPUT_BG); d.number_format = '0.000'; d.alignment = al()
    e = ws3.cell(row=i, column=5, value=f"=IF(D{i}=\"\",\"\",C{i}*D{i})")
    e.number_format = ILS; e.alignment = al()

for i in range(8, 28):
    ws3.cell(row=i, column=3).number_format = USD
    ws3.cell(row=i, column=3).alignment = al()
    d = ws3.cell(row=i, column=4)
    d.fill = fill(INPUT_BG); d.number_format = '0.000'; d.alignment = al()
    e = ws3.cell(row=i, column=5, value=f"=IF(D{i}=\"\",\"\",C{i}*D{i})")
    e.number_format = ILS; e.alignment = al()
    for col in [1, 2]: ws3.cell(row=i, column=col).alignment = al()

total_row3 = 28
ws3.merge_cells(f"A{total_row3}:D{total_row3}")
lbl(ws3, total_row3, 1, "סה\"כ תשלומים בפועל", bold=True, color=WHITE, bg=DARK)
val(ws3, total_row3, 5, f"=SUM(E4:E{total_row3-1})",
    fmt=ILS, bold=True, bg=DARK, color=WHITE)
ws3.row_dimensions[total_row3].height = 22

# ── Save ─────────────────────────────────────────────────
out = "Spinz_Import_Model.xlsx"
wb.save(out)
sys.stdout.write("Done: " + out + "\n")
